import re
import openpyxl
from .exceptions import InvalidExcelFileError
from .shipment import Shipment
from .base_module import ShipmentExractorModule


class ExcelModule(ShipmentExractorModule):

    def __init__(self, file):
        super().__init__(file)

    def extract_shipments(self):
        # Open the Workbook
        self.wb = openpyxl.load_workbook(self.file)
        # Create a list resulting of mapping the sheet's names into scores
        names = list(map(self.__to_score, self.wb.sheetnames))
        # Find biggest score
        max_score = max(names)

        if max_score == 1:
            raise InvalidExcelFileError(
                "The given .xlsx file doesn't have any valid data.")

        # Search the index of biggest score
        max_score_index = names.index(max_score)
        # Get the name of the sheet at the given index
        sheet_name = self.wb.sheetnames[max_score_index]
        # Get the sheet with it's name
        self.sheet = self.wb[sheet_name]
        # Get sheet's rows and columns data
        self.min_row = self.sheet.min_row
        self.max_row = self.sheet.max_row

        # Map firsts rows cell objects to their values
        first_row_values = list(
            map(lambda x: x.value, self.sheet[self.min_row]))

        first_row_are_headers = self.__is_first_row_the_header(
            first_row_values)

        self.columns_order = self.__default_columns_order()

        if first_row_are_headers:
            # Increase min row if first row are headers,
            # to avoid mapping first row to a Shipment object.
            self.min_row += 1
            # Only update if first row's values are headers,
            # if not, module asumes order of column is correct.
            self.columns_order = self.__update_columns_order(first_row_values)

        return self.__iter_and_get_shipments()

    def __to_score(self, name):
        """Use for mapping a sheet name into a score,
        given by the product of the max row count times
        the max column count.

        Args:
            name (str): to look for in workbook.

        Returns:
            int: the resulting score.
        """
        sheet = self.wb[name]
        score = sheet.max_row * sheet.max_column
        return score

    def __is_first_row_the_header(self, first_row_values: list) -> bool:
        """Returns wether the first row values correspond
        to the column's headers. If firsts rows cell values
        are purely strings, it will indicate first row are headers.

        Args:
            first_row_values (list): the workbook's first's row values

        Returns:
            bool: wether the first row is the header.
        """
        # Join values to a single temp string
        temp = "".join(first_row_values)
        # Regex search for numbers in temp
        regex_result = re.search(r'\d', temp)
        # Parse result to boolean if it contains numbers
        contains_numbers = bool(regex_result)
        if contains_numbers:
            # If first row contains numbers, return False,
            # because first row aren't headers at all.
            return False
        # If not, return True, because first row are headers.
        return True

    def __default_columns_order(self):
        return {
            'ID FLEX': 0, 'DOMICILIO': 1, 'ENTRECALLES': 2,
            'CODIGO POSTAL': 3, 'LOCALIDAD': 4, 'PARTIDO': 5,
            'DESTINATARIO': 6, 'DNI DESTINATARIO': 7,
            'TELEFONO DESTINATARIO': 8, 'DETALLE DEL ENVIO': 9,
            'CARGOS AL DESTINATARIO': 10,
        }

    def __update_columns_order(self, first_row_values: list) -> dict:
        columns_order = self.__default_columns_order()
        # Iterate over the first row's values and ther indexes

        # Default columns list
        columns = ['ID FLEX',
                   'DOMICILIO',
                   'ENTRECALLES',
                   'CODIGO POSTAL',
                   'LOCALIDAD',
                   'PARTIDO',
                   'DESTINATARIO',
                   'DNI DESTINATARIO',
                   'TELEFONO DESTINATARIO',
                   'DETALLE DEL ENVIO',
                   'CARGOS AL DESTINATARIO']
        for i, value in enumerate(first_row_values):
            # Increase index by one to get real column,
            # cause openpyxl index starts from 1
            col = i + 1
            # Update column header real index
            columns_order[value] = col
            # Removes item from defaults columns list
            columns.remove(value)

        # If any default column remains, give then -1 order in dict
        for column in columns:
            columns_order[column] = -1
        return columns_order

    def __iter_and_get_shipments(self) -> list:
        """Get the shipments from the worksheet.

        Returns:
            list[Shipment]: shipments found.
        """

        TRACKING = 'ID FLEX'
        DOMICILIO = 'DOMICILIO'
        ENTRECALLES = 'ENTRECALLES'
        COD_POSTAL = 'CODIGO POSTAL'
        LOCALIDAD = 'LOCALIDAD'
        PARTIDO = 'PARTIDO'
        DESTINATARIO = 'DESTINATARIO'
        DNI = 'DNI DESTINATARIO'
        TEL = 'TELEFONO DESTINATARIO'
        DETALLE = 'DETALLE DEL ENVIO'
        CARGOS = 'CARGOS AL DESTINATARIO'

        columns_order = self.columns_order

        shipments = []
        for row_i in range(self.min_row, self.max_row+1):
            shipment = Shipment()

            tracking = ""
            if columns_order[TRACKING] != -1:
                if self.sheet.cell(row=row_i, column=columns_order[TRACKING]
                                   ).value:
                    tracking = self.sheet.cell(
                        row=row_i, column=columns_order[TRACKING]).value
            shipment.tracking_id = tracking

            domicilio = ""
            if columns_order[DOMICILIO] != -1:
                if self.sheet.cell(row=row_i, column=columns_order[DOMICILIO]
                                   ).value:
                    domicilio = self.sheet.cell(
                        row=row_i, column=columns_order[DOMICILIO]).value
            shipment.domicilio = domicilio

            entrecalles = ""
            if columns_order[ENTRECALLES] != -1:
                if self.sheet.cell(row=row_i, column=columns_order[ENTRECALLES]
                                   ).value:
                    entrecalles = self.sheet.cell(
                        row=row_i, column=columns_order[ENTRECALLES]).value
            shipment.referencia = entrecalles

            cod_postal = ""
            if columns_order[COD_POSTAL] != -1:
                if self.sheet.cell(row=row_i, column=columns_order[COD_POSTAL]
                                   ).value:
                    cod_postal = self.sheet.cell(
                        row=row_i, column=columns_order[COD_POSTAL]).value
            shipment.codigo_postal = cod_postal

            localidad = ""
            if columns_order[LOCALIDAD] != -1:
                if self.sheet.cell(row=row_i, column=columns_order[LOCALIDAD]
                                   ).value:
                    localidad = self.sheet.cell(
                        row=row_i, column=columns_order[LOCALIDAD]).value
            shipment.localidad = localidad

            partido = ""
            if columns_order[PARTIDO] != -1:
                if self.sheet.cell(row=row_i, column=columns_order[PARTIDO]
                                   ).value:
                    partido = self.sheet.cell(
                        row=row_i, column=columns_order[PARTIDO]).value
            shipment.partido = partido

            dest = ""
            if columns_order[DESTINATARIO] != -1:
                if self.sheet.cell(
                    row=row_i, column=columns_order[DESTINATARIO]
                ).value:
                    dest = self.sheet.cell(
                        row=row_i, column=columns_order[DESTINATARIO]).value
            shipment.destinatario = dest

            dni = ""
            if columns_order[DNI] != -1:
                if self.sheet.cell(row=row_i, column=columns_order[DNI]).value:
                    dni = self.sheet.cell(
                        row=row_i, column=columns_order[DNI]).value
            shipment.dni_destinatario = dni

            tel = ""
            if columns_order[TEL] != -1:
                if self.sheet.cell(row=row_i, column=columns_order[TEL]).value:
                    tel = self.sheet.cell(
                        row=row_i, column=columns_order[TEL]).value
            shipment.phone = tel

            detalle = "0-1"
            if columns_order[DETALLE] != -1:
                if self.sheet.cell(row=row_i, column=columns_order[DETALLE]
                                   ).value:
                    detalle = self.sheet.cell(
                        row=row_i, column=columns_order[DETALLE]).value
            shipment.detalle_envio = detalle

            cargos = "0"
            if columns_order[CARGOS] != -1:
                if self.sheet.cell(row=row_i, column=columns_order[CARGOS]
                                   ).value:
                    cargos = self.sheet.cell(
                        row=row_i, column=columns_order[CARGOS]).value
            shipment.cargos = cargos

            shipment.clean()
            if shipment.is_not_empty():
                shipments.append(shipment)
        return shipments
