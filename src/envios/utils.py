from datetime import datetime, timedelta
from django.db.models.aggregates import Count
import calendar
from dateutil.relativedelta import relativedelta
from collections import Counter
from decimal import Decimal
from typing import Any, Callable, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import PatternFill
import unidecode
from account.models import Account
from clients.models import Client, Discount
from envios.models import BulkLoadEnvios, Envio
from tracking.models import TrackingMovement
from places.models import Partido, Town


ABBREVIATIONS = {
    "GRAL": "GENERAL",
    "GENERAL": "GRAL",
    "CNEL": "CORONEL",
    "CORONEL": "CNEL",
    "PTE": "PRESIDENTE",
    "PRESIDENTE": "PTE",
    "TTE": "TENIENTE",
    "TENIENTE": "TTE",
    "SGTO": "SARGENTO",
    "SARGENTO": "SGTO",
    "KM": "KILOMETRO",
    "KILOMETRO": "KM",
    "EST": "ESTACION",
    "ESTACION": "EST",
    "VICENTE": "VTE",
    "VTE": "VICENTE",
    "BS AS": "CABA",
    "BSAS": "CABA",
}


class NoSuggestionsAvailable(Exception):
    pass


def town_resolver(
    town: str, partido: str = None, zip_code: str = None
) -> Town:

    resolver = TownSuggestionResolver(town, partido, zip_code)
    result, reason = resolver.resolve()

    if not result:
        raise NoSuggestionsAvailable("Nothing found")

    return (result, reason)


class TownSuggestionResolver:

    BASE_REASON_STR = "Se encontró al buscar la localidad "

    def __init__(self, town: str, partido: str = None, zip_code: str = None):
        self.town_name: str = town.upper()
        self.zip_code_num: str = zip_code.upper() if zip_code else None
        self.partido_name: str = partido.upper() if partido else None
        self.found_towns: List[Town] = []
        self.result: Town = None
        self.next_step: Callable = self.__query_town_name
        self.reason_of_match: str = None

    def resolve(self) -> Tuple[Town, str]:
        self.__main_loop()
        return (self.result, self.reason_of_match)

    def __main_loop(self):
        while self.next_step:
            towns, self.next_step = self.next_step()
            if towns is not None and len(towns) > 0:
                if len(towns) > 1:
                    self.found_towns.extend(towns)
                else:
                    self.result = towns[0]
                    self.next_step = None

    def __update_reason(self, reason):
        self.reason_of_match = self.BASE_REASON_STR + reason

    def __query_town_name(self) -> Tuple[List[Town], Callable]:
        self.__update_reason("con el nombre")
        towns = list(Town.objects.filter(name=self.town_name))
        return (towns, self.__query_town_name_and_partido)

    def __query_town_name_and_partido(self) -> Tuple[List[Town], Callable]:
        self.__update_reason("con las letras del nombre y el partido")
        towns = list(
            Town.objects.filter(name__contains=self.town_name,
                                partido__name=self.partido_name)
        )
        return (towns, self.__query_unidecoded_town_name)

    def __query_unidecoded_town_name(self) -> Tuple[List[Town], Callable]:
        self.__update_reason("con el nombre limpio, sin comas, puntos, etc.")
        self.town_name = unidecode.unidecode(self.town_name).upper()
        towns = list(Town.objects.filter(name=self.town_name))
        return (towns, self.__query_no_enie_town_name)

    def __query_no_enie_town_name(self) -> Tuple[List[Town], Callable]:
        self.__update_reason("con el nombre limpio, sin comas, puntos, etc.")
        self.town_name = self.town_name.upper().replace("Ñ", "N")
        towns = list(Town.objects.filter(name=self.town_name))
        return (towns, self.__query_cleaned_town_name)

    def __query_cleaned_town_name(self) -> Tuple[List[Town], Callable]:
        self.__update_reason("con el nombre limpio, sin comas, puntos, etc.")
        self.town_name = self.town_name.replace(
            ".", "").replace("-", "").upper()
        towns = list(Town.objects.filter(name=self.town_name))
        return (towns, self.__query_town_part_of_name)

    def __query_town_part_of_name(self) -> Tuple[List[Town], Callable]:
        self.__update_reason(
            "con el nombre proporcionado como parte del nombre real.")
        towns = list(Town.objects.filter(name__contains=self.town_name))
        return (towns, self.__query_replacing_abbreviations)

    def __query_replacing_abbreviations(self) -> Tuple[List[Town], Callable]:
        self.__update_reason("con el nombre sin abreviaturas.")
        town = self.town_name
        for key, value in ABBREVIATIONS.items():
            if key in town:
                town.replace(key, value)
        towns = list(Town.objects.filter(name=town))
        return (towns, self.__query_postal_code)

    def __query_postal_code(self) -> Tuple[List[Town], Callable]:
        self.__update_reason("con el código postal.")
        towns = None
        if self.zip_code_num:
            towns = list(Town.objects.filter(zipcode__code=self.zip_code_num))
        return (towns, self.__query_most_matching_name)

    def __query_most_matching_name(self) -> Tuple[List[Town], Callable]:
        self.__update_reason("que tuvo más coincidencias.")
        town = self.town_name
        words = filter(lambda w: len(w) > 3, town.split(" "))
        towns = []
        for word in words:
            towns.extend(list(Town.objects.filter(name__contains=word)))
        if towns:
            counted = Counter(towns)
            ordered = counted.most_common()
            most_common = counted.most_common(1)[0][1]
            towns.clear()
            for item, score in ordered:
                if score == most_common:
                    towns.append(item)
                else:
                    break
        return (towns, self.__query_found_towns_most_matching_name)

    def __query_found_towns_most_matching_name(
            self
    ) -> Tuple[List[Town], Callable]:
        self.__update_reason("que tuvo más coincidencias.")
        if self.found_towns:
            counted = Counter(self.found_towns)
            ordered = counted.most_common()
            most_common = counted.most_common(1)[0][1]
            self.found_towns.clear()
            for item, score in ordered:
                if score == most_common:
                    self.found_towns.append(item)
                else:
                    break
        return (self.found_towns, self.__query_partido_as_name)

    def __query_partido_as_name(self) -> Tuple[List[Town], Callable]:
        self.__update_reason("con el partido como nombre.")
        towns = None
        if self.partido_name:
            query_name = self.partido_name.upper()
            towns = list(Town.objects.filter(name=query_name))
        return (towns, self.__query_partido_as_name_replacing_abbreviations)

    def __query_partido_as_name_replacing_abbreviations(
        self
    ) -> Tuple[List[Town], Callable]:
        self.__update_reason("con el partido como nombre sin abreviaturas.")
        towns = None
        if self.partido_name:
            query_name = self.partido_name.upper()
            for key, value in ABBREVIATIONS.items():
                if key in query_name:
                    query_name.replace(key, value)
            towns = list(Town.objects.filter(name=query_name))
        return (towns, self.__query_partido_as_part_of_name)

    def __query_partido_as_part_of_name(self) -> Tuple[List[Town], Callable]:
        self.__update_reason("con el partido como parte nombre.")
        towns = None
        if self.partido_name:
            query_name = self.partido_name.upper()
            towns = list(Town.objects.filter(name__contains=query_name))
        return (towns, self.__query_first_town_in_partido)

    def __query_first_town_in_partido(self):
        self.reason_of_match = "Es la primera localidad \
            del partido proporcionado"
        towns = []
        if self.partido_name:
            partidos = Partido.objects.filter(name=self.partido_name)
            if partidos:
                towns = partidos[0].town_set.all()
                if towns:
                    towns.order_by("name")
        return (towns, None)


def create_xlsx_workbook(
    csv_str: str, cells_to_paint: str = None
) -> Workbook:
    # Create an excel workbook
    wb = Workbook()
    # Get first sheet (1 is created when wb created)
    sheet = wb.active
    # Change title to 'Datos'
    sheet.title = 'Datos'
    # Get the sheet recently changed
    sheet = wb.get_sheet_by_name('Datos')
    # The default amount of columns
    COLUMNS = 10
    # Parse the csv string to list of lists
    result = [row.split(",") for row in csv_str.split("\n")]
    col_names = [
        "ID FLEX", "DOMICILIO", "ENTRECALLES", "CODIGO POSTAL",
        "LOCALIDAD", "PARTIDO", "DESTINATARIO", "DNI DESTINATARIO",
        "TELEFONO DESTINATARIO", "DETALLE DEL ENVIO"]
    result[0] = col_names
    # Iterate over rows and cols indexes
    for i in range(len(result)):
        for j in range(COLUMNS):

            # Get the cell at i and j
            cell = sheet.cell(row=i+1, column=j+1)

            # Set the value to the corresponding csv row and col
            if i != 0 and j == 4:
                try:
                    if result[i][j] and str(result[i][j]).isdigit():
                        cell.value = Town.objects.get(id=result[i][j]).name
                    else:
                        cell.value = ""
                except Town.DoesNotExist:
                    cell.value = ""
            else:
                cell.value = result[i][j]

            # Cell reference as pair values as string
            cell_ref = f"{i},{j}"

            # If the cell_ref is in csv_errors
            if cells_to_paint and cell_ref in cells_to_paint:
                # Change the bg color to yellow
                cell.fill = PatternFill("solid", fgColor="FFFF00")
    return wb


def create_empty_xlsx_workbook() -> Workbook:
    # Create an excel workbook
    wb = Workbook()
    # Get first sheet (1 is created when wb created)
    sheet = wb.active
    # Change title to 'Datos'
    sheet.title = 'Datos'
    # Get the sheet recently changed
    sheet = wb.get_sheet_by_name('Datos')
    # The default amount of columns
    COLUMNS = 10
    col_names = [
        "ID FLEX", "DOMICILIO", "ENTRECALLES", "CODIGO POSTAL",
        "LOCALIDAD", "PARTIDO", "DESTINATARIO", "DNI DESTINATARIO",
        "TELEFONO DESTINATARIO", "DETALLE DEL ENVIO"]
    for i in range(COLUMNS):
        cell = sheet.cell(row=1, column=i+1)
        cell.value = col_names[i]
    return wb


def bulk_create_envios(
    bulk_load_envios: BulkLoadEnvios
) -> Tuple[List[Envio], List[str]]:
    """
    Create envios from the csv file and return them.
    Ids are returned just because PostgreSQL supports it.
    """
    # Get the csv file
    envios = []
    unused_flex_ids = []
    for i, row in enumerate(bulk_load_envios.csv_result.split("\n")):
        if i == 0 or "traking_id" in row:
            continue
        cols = row.split(",")
        kwargs = __cols_to_kwargs(cols, bulk_load_envios)
        if 'is_flex' in kwargs and kwargs['is_flex'] and kwargs['flex_id']:
            if Envio.objects.filter(flex_id=kwargs['flex_id']).exists():
                unused_shipment = "'{fid}: {street}, {town}'".format(
                    fid=kwargs['flex_id'],
                    street=kwargs['street'],
                    town=kwargs['town']
                )
                unused_flex_ids.append(unused_shipment)
                continue
        envio = Envio(**kwargs)
        envio.deposit = bulk_load_envios.deposit
        envio.save()
        envios.append(envio)

    if len(envios) > 0:
        author = envios[0].updated_by
        tm = TrackingMovement(
            created_by=author,
            action=TrackingMovement.ACTION_ADDED_TO_SYSTEM,
            result=TrackingMovement.RESULT_ADDED_TO_SYSTEM,
            to_deposit=bulk_load_envios.deposit
        )
        tm.save()
        tm.envios.add(*envios)
    return (envios, unused_flex_ids)


def convert_date(date_string: str):
    current_year = datetime.now().year
    day, month = date_string.split('-')

    months_dict = {
        'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
        'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12
    }
    month_number = months_dict[month.lower()]

    converted_date = datetime(current_year, month_number, int(day))

    min_date = datetime.now() - timedelta(days=10)
    if converted_date < min_date:
        converted_date = converted_date.replace(year=current_year + 1)

    return converted_date


def __cols_to_kwargs(
    cols: List[str], bulk_load_envios: BulkLoadEnvios
) -> Dict[str, Any]:
    kwargs = {
        'street': cols[1],
        'remarks': cols[2],
        'zipcode': cols[3],
        'town': Town.objects.get(pk=cols[4]),
        'name': cols[6],
        'doc': cols[7],
        'phone': cols[8],
        'detail': cols[9] if cols[9] else "0-1",
        'updated_by': bulk_load_envios.created_by,
        'client': bulk_load_envios.client,
        'bulk_upload_id': bulk_load_envios.pk,
    }

    if cols[10] != "":
        kwargs['max_delivery_date'] = convert_date(cols[10])

    if cols[0] != '' and not cols[0].isspace():
        kwargs['is_flex'] = True
        kwargs['flex_id'] = cols[0]

    return kwargs


DETAIL_CODES = {
    "0": 'max_5k_price',
    "1": 'bulto_max_10k_price',
    "2": 'bulto_max_20k_price',
    "3": 'miniflete_price',
    "4": 'tramite_price',
    "5": 'camioneta_price',
}

VERBOSE_NAMES_FOR_DETAIL_CODES = {
    "0": 'Paquete(s) hasta 5kg',
    "1": 'Bulto(s) hasta 10kg',
    "2": 'Bulto(s) hasta 20kg',
    "3": 'miniflete(s)',
    "4": 'trámite(s)',
    "5": 'camioneta(s) completa(s)',
}


def calculate_price(envio: Envio):
    """
    Performs the calculation for the price of the 'envio'.
    It applies the discount if it is available.
    The operations are made with Decimal to avoid the rounding errors.

    Returns:
        Decimal: The price of the 'envio'.
    """
    # Get the town of recipient's address
    town = envio.town

    # #  Get the code. If 'envio' is from flex, return flex's code for
    # # given town, else the normal code.
    # price_code = town.flex_code if envio.is_flex else town.delivery_code

    # Initialize total price to 0
    total_price = Decimal(0)

    if envio.is_flex:
        # Get the flex price
        total_price = Decimal(str(town.flex_code.price))
    else:
        delivery_code = town.delivery_code
        # Get the detail for the given envio
        detail_codes = envio.detail.split(',')
        # Get the price for each package detail
        for detail_code in detail_codes:
            # Unpack code and amount spliting by '-'
            code, amount = detail_code.split('-')
            # Get multiplier for the given package detail
            attr = DETAIL_CODES[code]
            multiplier = getattr(delivery_code, attr)
            # Multiply the price by the multiplier for given package
            # detail, times the amount of packages
            result = Decimal(str(multiplier)) * Decimal(amount)
            # Add the result to the total price
            total_price += result

    # Get single discount for envio's client and partido, if and
    # only if the envio and discount match the is_for_flex and
    # is_flex flags
    discount = Discount.objects.filter(
        client=envio.client,
        is_for_flex=envio.is_flex,
        partidos__in=[envio.town.partido]
    ).first()

    # If discount exists, apply it to the total price
    if discount:
        discount = Decimal(discount.amount) / Decimal(100)
        total_discount = total_price * discount
        result = total_price - total_discount
        return result

    # return the total price
    return total_price


def get_detail_readable(envio: Envio):
    if envio.is_flex and envio.flex_id is not None:
        return f'Paquete Flex #{envio.flex_id}'
    details = map(map_detail, envio.detail.split(","))
    return ", ".join(details)


def map_detail(detail):
    parts = detail.split("-")
    index = parts[0]
    amount = parts[1]
    if index in DETAIL_CODES.keys():
        verbose_name = VERBOSE_NAMES_FOR_DETAIL_CODES[index]
        return f'{amount}x {verbose_name}'.lower()
    return ""


MONTHS = {
    '01': 'Enero',
    '02': 'Febrero',
    '03': 'Marzo',
    '04': 'Abril',
    '05': 'Mayo',
    '06': 'Junio',
    '07': 'Julio',
    '08': 'Agosto',
    '09': 'Septiembre',
    '10': 'Octubre',
    '11': 'Noviembre',
    '12': 'Diciembre'
}


def get_stats_4_last_12_months() -> Tuple[List[str], List[int]]:
    """
    Returns a Tuple containing two lists, one with last 12 months'
    name labels, and the other with corresponding delivered shipments count.

    Returns:
        Tuple[List[str], List[int]]: The Tuple containing both lists.
    """
    now = datetime.now()
    months = []
    counts = []
    for _ in range(12):
        months.insert(0, MONTHS[now.strftime('%m')])
        last_day_in_month = calendar.monthrange(now.year, now.month)[1]
        from_date = datetime(now.year, now.month, 1, 0, 0, 0)
        to_date = datetime(now.year, now.month, last_day_in_month, 23, 59, 59)
        delivered_count = TrackingMovement.objects.filter(
            date_created__range=(from_date, to_date),
            result=TrackingMovement.RESULT_DELIVERED
        ).count()
        counts.insert(0, delivered_count)
        now = now - relativedelta(months=1)
    return months, counts


def get_stats_daily_deliveries():
    delivered = Envio.objects.filter(status=Envio.STATUS_DELIVERED)
    sdate = (
        delivered.order_by('date_delivered')
        .first().date_delivered)
    edate = delivered.order_by('-date_delivered').first().date_delivered

    range_of_days = range((edate - sdate).days + 2)
    date_list = [sdate+timedelta(days=x) for x in range_of_days]

    qdelivery = "date_delivered__date"
    qcount = "envio_count"
    annotations = {qcount: Count('id')}

    result = (
        Envio.objects
        .filter(status=Envio.STATUS_DELIVERED)
        .values(qdelivery)
        .annotate(**annotations)
        .order_by(qdelivery)
    )

    date_fmt = "%d-%m-%y"

    def to_str(date): return date.strftime(date_fmt)

    count_dict = {to_str(entry[qdelivery]): entry[qcount] for entry in result}

    labels, counts = [], []
    for i, date in enumerate(date_list):
        date_str = to_str(date)
        label = date_str if (date.day == 1 or i in [0, len(date_list)]) else ""
        labels.append(label)
        count = count_dict.get(date_str, 0)
        counts.append(count)

    return labels, counts


def get_client_share() -> Tuple[Tuple[str], Tuple[int]]:
    """
    Returns a Tuple containing two tuples, one with clients'
    names and the other with matching shipments count, both
    ordered by client shipments count descending.

    Returns:
        Tuple[Tuple[str], Tuple[int]]: The Tuple containing both tuples.
    """
    clients = []
    envios_counts = []
    clients_qs = Client.objects.all()\
        .values('name')\
        .annotate(Count('envio'))\
        .order_by('-envio__count')\
        .values_list('name', 'envio__count')
    others = 0
    for name, envio_count in clients_qs:
        if clients_qs.count() > 0:
            max_count = clients_qs.first()[1]
            ten_percent = 1 * max_count / 100
            if envio_count > ten_percent:
                clients.append(name)
                envios_counts.append(envio_count)
            else:
                others += envio_count
    clients.append('Otros')
    envios_counts.append(others)
    return clients, envios_counts


def get_carrier_share() -> Tuple[Tuple[str], Tuple[int]]:
    """
    Returns a Tuple containing two tuples, one with carriers'
    names and the other with matching shipments count, both
    ordered by client shipments count descending.

    Returns:
        Tuple[Tuple[str], Tuple[int]]: The Tuple containing both tuples.
    """
    carriers = []
    envios_counts = []
    carriers_qs = Account.objects\
        .filter(envios_delivered_by__gt=0)\
        .values('username', 'first_name', 'last_name')\
        .annotate(envio_count=Count('envios_delivered_by'))\
        .order_by('-envio_count')\
        .values_list('username', 'first_name', 'last_name', 'envio_count')

    others = 0
    if carriers_qs.count() > 0:
        max_count = carriers_qs.first()[-1]
        for username, first_name, last_name, envio_count in carriers_qs:
            name = f'{first_name} {last_name} ({username})'
            ten_percent = 1 * max_count / 100
            if envio_count > ten_percent:
                carriers.append(name)
                envios_counts.append(envio_count)
            else:
                others += envio_count

    carriers.append('Otros')
    envios_counts.append(others)
    return carriers, envios_counts
